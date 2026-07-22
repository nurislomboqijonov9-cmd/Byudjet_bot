"""Mijoz hisobotini Excel (.xlsx) qilib beradi."""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


BRAND = "1F4B45"
LIGHT = "EEF2F0"
REDBG = "F6E1DA"
REDINK = "B24A31"
_thin = Side(style="thin", color="D9DFDD")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _dmy(s):
    s = str(s or "")[:10]
    p = s.split("-")
    return f"{p[2]}.{p[1]}.{p[0][2:]}" if len(p) == 3 else s


def _som(n):
    return f"{round(n or 0):,}".replace(",", " ")


def _status_uz(s):
    return {"faol": "Faol", "nofaol": "Nofaol", "sotuv": "Sotuv"}.get(s, "-")


def _banner(ws, text, row, span=8, fill=BRAND, color="FFFFFF", size=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=color, size=size)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22


def _header_row(ws, row, heads):
    for i, h in enumerate(heads, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=LIGHT)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")


def mijoz_excel(d):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hisobot"
    ws.column_dimensions["A"].width = 6
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["I"].width = 22
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["K"].width = 14

    def title(text, row, span=8, fill=BRAND, color="FFFFFF", size=13):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color=color, size=size)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22

    r = 1
    title(f"MIJOZ: {d['mijoz']}", r); r += 1
    for lbl, val in [("Telefon", d.get("telefon") or "-"), ("Manzil", d.get("adres") or "-"),
                     ("Status", _status_uz(d.get("status")))]:
        ws.cell(row=r, column=1, value=lbl).font = Font(bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.cell(row=r, column=2, value=val)
        r += 1
    r += 1

    # Partiyalar jadvali
    title("PARTIYALAR (chiqgan mollar)", r, span=11); r += 1
    heads = ["№", "Mahsulot", "Jami", "Qolgan", "Kunlik narx", "Chiqgan sana", "Kun", "Summa (so'm)", "Manzil"]
    for i, h in enumerate(heads, 1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=LIGHT)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    r += 1
    for p in d.get("partiyalar", []):
        row = [p["partiya_raqam"], p["mahsulot"], p["miqdor"], p["qolgan"],
               p["kunlik_narx"], _dmy(p["chiqgan_sana"]), p["kunlar"], round(p["narx"]),
               p.get("manzil") or "—"]
        for i, v in enumerate(row, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER
            if i in (3, 4, 5, 7, 8):
                c.alignment = Alignment(horizontal="right")
        r += 1
    r += 1

    # Manzillar bo'yicha (qolgan tovarlar qaysi manzilda)
    manzillar = d.get("manzillar") or []
    if any(m.get("manzil") != "Manzil belgilanmagan" for m in manzillar):
        title("MANZILLAR BO'YICHA (qolgan)", r, span=11); r += 1
        for m in manzillar:
            ws.cell(row=r, column=1, value=f"📍 {m['manzil']}").font = Font(bold=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            c = ws.cell(row=r, column=7, value=f"{_som(m['qolgan_dona'])} dona")
            c.font = Font(bold=True); c.alignment = Alignment(horizontal="right")
            for col in range(1, 12):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT)
            r += 1
            for it in m["items"]:
                ws.cell(row=r, column=2, value=it["mahsulot"]).border = BORDER
                cc = ws.cell(row=r, column=3, value=it["qolgan"]); cc.border = BORDER
                cc.alignment = Alignment(horizontal="right")
                cs = ws.cell(row=r, column=8, value=round(it["narx"])); cs.border = BORDER
                cs.alignment = Alignment(horizontal="right")
                r += 1
        r += 1

    # Qaytarishlar
    rets = [(p, rr) for p in d.get("partiyalar", []) for rr in p.get("qaytarishlar", [])]
    if rets:
        title("QAYTARISHLAR", r); r += 1
        for h, col in [("Partiya", 1), ("Mahsulot", 2), ("Soni", 3), ("Sana", 4)]:
            c = ws.cell(row=r, column=col, value=h)
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor=LIGHT); c.border = BORDER
        r += 1
        for p, rr in rets:
            for i, v in enumerate([p["partiya_raqam"], p["mahsulot"], rr["miqdor"], _dmy(rr["qaytgan_sana"])], 1):
                ws.cell(row=r, column=i, value=v).border = BORDER
            r += 1
        r += 1

    # To'lovlar
    if d.get("tolovlar"):
        title("TO'LOVLAR (predoplata)", r); r += 1
        for p in d["tolovlar"]:
            ws.cell(row=r, column=1, value=_dmy(p["sana"]))
            ws.cell(row=r, column=2, value=(p.get("izoh") or "to'lov"))
            ws.cell(row=r, column=3, value=round(p["summa"])).alignment = Alignment(horizontal="right")
            r += 1
        r += 1

    # Qo'shimcha (yo'lkira / remont)
    if d.get("qoshimcha"):
        title("QO'SHIMCHA (yo'lkira / remont)", r); r += 1
        for q in d["qoshimcha"]:
            ws.cell(row=r, column=1, value=_dmy(q["sana"]))
            ws.cell(row=r, column=2, value=("Yo'lkira" if q["tur"] == "yolkira" else "Remont"))
            ws.cell(row=r, column=3, value=round(q["summa"])).alignment = Alignment(horizontal="right")
            r += 1
        r += 1

    # Yakuniy hisob
    title("YAKUNIY HISOB", r); r += 1
    rows = [("Hisoblangan (ijara)", d.get("hisoblangan", 0)),
            ("Yo'lkira", d.get("yolkira", 0)),
            ("Remont", d.get("remont", 0)),
            ("To'langan", d.get("tolangan", 0)),
            ("QOLGAN QARZ", d.get("qolgan_qarz", 0))]
    for lbl, val in rows:
        bold = lbl == "QOLGAN QARZ"
        c1 = ws.cell(row=r, column=1, value=lbl)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c1.font = Font(bold=bold)
        c3 = ws.cell(row=r, column=3, value=f"{_som(val)} so'm")
        c3.font = Font(bold=bold)
        c3.alignment = Alignment(horizontal="right")
        if bold:
            for col in (1, 2, 3):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT)
        r += 1

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def umumiy_excel(mlist, sana=None):
    """Butun mijozlar bazasi bitta jadvalda (db.mijozlar() natijasi)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Umumiy"
    for c, w in {"A": 5, "B": 24, "C": 16, "D": 9, "E": 11, "F": 15, "G": 14, "H": 16}.items():
        ws.column_dimensions[c].width = w

    r = 1
    sarlavha = "UMUMIY MIJOZLAR RO'YXATI"
    if sana:
        sarlavha += f" ({_dmy(sana)})"
    _banner(ws, sarlavha, r, span=8); r += 1
    heads = ["№", "Mijoz", "Telefon", "Status", "Qolgan dona", "Hisoblangan", "To'langan", "Qolgan qarz"]
    _header_row(ws, r, heads); r += 1

    n = 0
    t_his = t_tol = t_qarz = 0.0
    for m in mlist:
        n += 1
        t_his += m.get("jami", 0) or 0
        t_tol += m.get("tolangan", 0) or 0
        t_qarz += m.get("qolgan_qarz", 0) or 0
        vals = [n, m.get("mijoz", ""), m.get("telefon") or "-", _status_uz(m.get("status")),
                m.get("jami_qolgan", 0), round(m.get("jami", 0) or 0),
                round(m.get("tolangan", 0) or 0), round(m.get("qolgan_qarz", 0) or 0)]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER
            if i in (5, 6, 7, 8):
                c.alignment = Alignment(horizontal="right")
        r += 1

    # Jami qatori
    ws.cell(row=r, column=1, value="JAMI").font = Font(bold=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    for i, v in [(6, t_his), (7, t_tol), (8, t_qarz)]:
        c = ws.cell(row=r, column=i, value=round(v))
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="right")
    for col in range(1, 9):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def qarzdorlar_excel(qlist, limit_kun, sana=None):
    """Qarzdorlar ro'yxati (db.qarzdorlar() natijasi). Chegaradan oshganlar qizil belgilanadi."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Qarzdorlar"
    for c, w in {"A": 5, "B": 14, "C": 24, "D": 16, "E": 16, "F": 14, "G": 13, "H": 10}.items():
        ws.column_dimensions[c].width = w

    r = 1
    sarlavha = f"QARZDORLAR RO'YXATI · chegara {limit_kun} kunlik ijara"
    if sana:
        sarlavha += f" ({_dmy(sana)})"
    _banner(ws, sarlavha, r, span=8); r += 1
    heads = ["№", "Holat", "Mijoz", "Telefon", "Qarz (so'm)", "Kunlik ijara", "Necha kunlik", "Status"]
    _header_row(ws, r, heads); r += 1

    n = 0
    t_qarz = 0.0
    for x in qlist:
        n += 1
        t_qarz += x.get("qarz", 0) or 0
        over = x.get("over")
        holat = "Yig'ish kerak" if over else "Kuzatuvda"
        kun = "rental yo'q" if x.get("kun") is None else round(x["kun"])
        vals = [n, holat, x.get("ism", ""), x.get("telefon") or "-",
                round(x.get("qarz", 0) or 0), round(x.get("rate", 0) or 0),
                kun, _status_uz(x.get("status"))]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER
            if i in (5, 6, 7):
                c.alignment = Alignment(horizontal="right")
            if over:
                c.fill = PatternFill("solid", fgColor=REDBG)
                if i == 2:
                    c.font = Font(bold=True, color=REDINK)
        r += 1

    # Jami qatori
    ws.cell(row=r, column=1, value="JAMI").font = Font(bold=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(row=r, column=5, value=round(t_qarz))
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="right")
    for col in range(1, 9):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def brov_excel(groups, sana=None):
    """Brovdan olinganlar: kimdan qaysi tovardan qancha, qancha qaytarilgan."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Brovdan"
    for c, w in {"A": 5, "B": 22, "C": 22, "D": 12, "E": 12, "F": 12, "G": 14, "H": 22}.items():
        ws.column_dimensions[c].width = w
    r = 1
    sarl = "BROVDAN OLINGANLAR (boshqadan olib turilgan)"
    if sana:
        sarl += f" · {_dmy(sana)}"
    _banner(ws, sarl, r, span=8); r += 1
    _header_row(ws, r, ["№", "Kimdan", "Mahsulot", "Olingan", "Qaytarilgan", "Qolgan", "Sana", "Izoh"]); r += 1
    n = 0
    t_olindi = t_qolgan = 0.0
    for g in groups:
        for b in g["items"]:
            n += 1
            t_olindi += float(b["miqdor"])
            t_qolgan += float(b["qolgan"])
            vals = [n, g["kim"], b["mahsulot"], b["miqdor"], b["qaytgan"], b["qolgan"],
                    _dmy(b["sana"]), b.get("izoh") or "—"]
            for i, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=i, value=v)
                c.border = BORDER
                if i in (4, 5, 6):
                    c.alignment = Alignment(horizontal="right")
                if i == 6 and float(b["qolgan"]) > 0:
                    c.fill = PatternFill("solid", fgColor=REDBG)
                    c.font = Font(bold=True, color=REDINK)
            r += 1
    ws.cell(row=r, column=1, value="JAMI").font = Font(bold=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    for i, v in [(4, t_olindi), (6, t_qolgan)]:
        c = ws.cell(row=r, column=i, value=round(v, 2))
        c.font = Font(bold=True); c.alignment = Alignment(horizontal="right")
    for col in range(1, 9):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
